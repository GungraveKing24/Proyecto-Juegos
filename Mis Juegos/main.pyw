from PySide6.QtWidgets import *
import sys 
from form import *
from PySide6 import *
import pandas as pd
from tkinter import *
from PyQt5 import QtWidgets
from openpyxl import load_workbook



class Juegos(QMainWindow): # a esto le puedo camiar el nombre de la clase nada más
    def __init__(self):
        super().__init__()
        self.ui = Ui_Dialog() #permitir utilizar todos los objetos del form
        self.ui.setupUi(self)
        #Botones para cada metodo
        #self.ui.pushButton.clicked.connect(self.CargarDatos)
        self.ui.pushButton_2.clicked.connect(self.CargarDatos)
        self.ui.pushButton.clicked.connect(self.ingresardatos)

    def CargarDatos(self):
        df = pd.read_excel('Lista de Juegos.xlsx')
        if df.size == 0:
            return
        
        df.fillna('', inplace=True)
        self.ui.table.setRowCount(df.shape[0])
        self.ui.table.setColumnCount(df.shape[1])
        self.ui.table.setHorizontalHeaderLabels(df.columns)

        for row in df.iterrows():
            values = row[1]
            for col_index, value in enumerate(values):
                if isinstance(value, (float, int)):
                    value = '{0:0,.0f}'.format(value)
                tableitem = QTableWidgetItem(str(value))
                self.ui.table.setItem(row[0], col_index, tableitem)

    
    def ingresardatos(self):
        pedirjuegos = self.ui.lineEdit.text()
        pedirestado= self.ui.comboBox.currentText()
        pedirdatosextra = self.ui.textEdit.toPlainText()
        
        workbook = load_workbook('Lista de Juegos.xlsx')
        hoja = workbook['Juegos']

        ultimafila = hoja.max_row - 1
        hoja.append([pedirjuegos, pedirestado, pedirdatosextra])
        workbook.save('Lista de Juegos.xlsx')
 

if __name__=="__main__":
    app=QApplication(sys.argv)
    myapp=Juegos()
    myapp.show()
    sys.exit(app.exec_())